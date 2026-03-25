#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import shutil
import sys
import tempfile
import zipfile
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import extract_report_csv as extractor


@dataclass
class PdfExtract:
    pdf_path: Path
    report_date: dt.date | None
    amounts: dict[int, int]


_ALLOWED_FILENAME_CHAR = re.compile(r"[А-Яа-яЁёA-Za-z0-9 _.,()\-+/\\]")


def _text_quality_score(s: str) -> int:
    good = 0
    bad = 0
    cyr = 0
    for ch in s:
        if _ALLOWED_FILENAME_CHAR.fullmatch(ch):
            good += 1
        else:
            bad += 1
        if ("А" <= ch <= "я") or ch in "Ёё":
            cyr += 1
    return good + cyr * 2 - bad * 3


def _decode_zip_member_name(raw_name: str, flag_bits: int) -> str:
    # Bit 11 means UTF-8 filename in ZIP spec; keep as is.
    if flag_bits & 0x800:
        return raw_name

    try:
        raw_bytes = raw_name.encode("cp437")
    except UnicodeEncodeError:
        return raw_name

    candidates = [raw_name]
    for enc in ("cp866", "cp1251", "utf-8"):
        try:
            candidates.append(raw_bytes.decode(enc))
        except UnicodeDecodeError:
            continue

    return max(candidates, key=_text_quality_score)


def _repair_mojibake_name(text: str) -> str:
    # For already-extracted folders with broken names, try a conservative repair.
    try:
        raw = text.encode("cp437")
    except UnicodeEncodeError:
        return text

    candidates = [text]
    for enc in ("cp866", "cp1251"):
        try:
            candidates.append(raw.decode(enc))
        except UnicodeDecodeError:
            continue

    scored = [(c, _text_quality_score(c)) for c in candidates]
    best, best_score = max(scored, key=lambda x: x[1])
    current_score = scored[0][1]

    if best != text and best_score >= current_score + 3:
        return best
    return text


def extract_zip_preserve_ru_names(zip_path: Path, dst_root: Path) -> None:
    dst_root.mkdir(parents=True, exist_ok=True)
    base = dst_root.resolve()

    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            decoded_name = _decode_zip_member_name(info.filename, info.flag_bits)
            decoded_name = decoded_name.replace("\\", "/")
            rel = Path(decoded_name)
            target = (dst_root / rel).resolve()

            # Basic zip-slip guard.
            if target != base and base not in target.parents:
                print(f"WARN zip member skipped (unsafe path): {decoded_name}", file=sys.stderr)
                continue

            if info.is_dir() or decoded_name.endswith("/"):
                target.mkdir(parents=True, exist_ok=True)
                continue

            target.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(info, "r") as src, target.open("wb") as out:
                shutil.copyfileobj(src, out)


def parse_iso_date(s: str | None) -> dt.date | None:
    if not s:
        return None
    try:
        return dt.date.fromisoformat(s)
    except ValueError:
        return None


def parse_date_from_name(name: str) -> dt.date | None:
    patterns = [
        r"(?P<d>\d{2})[._\-](?P<m>\d{2})[._\-](?P<y>\d{4})",
        r"(?P<y>\d{4})[._\-](?P<m>\d{2})[._\-](?P<d>\d{2})",
        r"(?P<d>\d{2})(?P<m>\d{2})(?P<y>\d{4})",
    ]
    for pat in patterns:
        m = re.search(pat, name)
        if not m:
            continue
        try:
            return dt.date(int(m.group("y")), int(m.group("m")), int(m.group("d")))
        except ValueError:
            continue
    return None


def parse_period_header_date(text: str) -> dt.date | None:
    if not text:
        return None

    s = text.lower().replace("ё", "е")
    s = re.sub(r"\s+", " ", s)

    patterns: list[tuple[str, tuple[int, int]]] = [
        (r"за\s*1\s*(?:квартал|кв\.?)\s*(20\d{2})", (3, 31)),
        (r"за\s*1\s*полугодие\s*(20\d{2})", (6, 30)),
        (r"за\s*9\s*месяц(?:ев|а)?\s*(20\d{2})", (9, 30)),
        (r"за\s*(20\d{2})\s*г(?:од|ода)?", (12, 31)),
    ]

    for pat, (month, day) in patterns:
        m = re.search(pat, s, flags=re.IGNORECASE)
        if not m:
            continue
        try:
            year = int(m.group(1))
            return dt.date(year, month, day)
        except ValueError:
            continue

    return None


def normalize_code(value: object) -> int | None:
    if value is None:
        return None
    s = re.sub(r"\D", "", str(value))
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def extract_single_pdf(pdf_path: Path, dpi: int, lang: str, psm: int) -> PdfExtract:
    ocr_text = extractor.ocr_pdf(pdf_path, dpi=dpi, lang=lang, psm=psm)
    ocr_data = extractor.extract_code_amounts(ocr_text)

    text_layer = ""
    text_data: dict[int, int] = {}
    try:
        text_layer = extractor.extract_text_from_pdf(pdf_path)
        text_data = extractor.extract_code_amounts(text_layer)
    except RuntimeError:
        pass

    merged = dict(text_data)
    merged.update(ocr_data)
    merged = extractor.fill_derived_values(merged)

    date_iso = extractor.extract_reporting_date(text_layer) or extractor.extract_reporting_date(ocr_text)
    report_date = (
        parse_iso_date(date_iso)
        or parse_date_from_name(pdf_path.name)
        or parse_period_header_date(text_layer)
        or parse_period_header_date(ocr_text)
        or parse_period_header_date(pdf_path.name)
    )

    return PdfExtract(pdf_path=pdf_path, report_date=report_date, amounts=merged)


def merge_amounts(base: dict[int, int], extra: dict[int, int]) -> dict[int, int]:
    out = dict(base)
    for k, v in extra.items():
        if k not in out:
            out[k] = v
            continue
        if out[k] != v:
            # Conservative conflict rule: keep value with larger absolute magnitude.
            if abs(v) > abs(out[k]):
                out[k] = v
    return out


def group_pdfs_by_date(
    pdf_files: Iterable[Path],
    dpi: int,
    lang: str,
    psm: int,
    infer_pair_date: bool = False,
) -> dict[dt.date, dict[int, int]]:
    grouped: dict[dt.date, dict[int, int]] = {}
    extracted: list[PdfExtract] = []

    for pdf in sorted(pdf_files):
        ex = extract_single_pdf(pdf, dpi=dpi, lang=lang, psm=psm)
        if not ex.amounts:
            print(f"WARN {pdf.name}: no extracted code/value pairs", file=sys.stderr)
            continue
        extracted.append(ex)

    if infer_pair_date and len(extracted) == 2:
        known_dates = {ex.report_date for ex in extracted if ex.report_date is not None}
        undated_items = [ex for ex in extracted if ex.report_date is None]
        if len(known_dates) == 1 and len(undated_items) == 1:
            inferred_date = next(iter(known_dates))
            undated_items[0].report_date = inferred_date
            print(
                f"WARN {undated_items[0].pdf_path.name}: report date inferred from paired PDF -> {inferred_date.isoformat()}",
                file=sys.stderr,
            )

    undated = 0
    for ex in extracted:
        if ex.report_date is None:
            undated += 1
            print(f"WARN {ex.pdf_path.name}: report date not detected; file skipped", file=sys.stderr)
            continue

        grouped.setdefault(ex.report_date, {})
        grouped[ex.report_date] = merge_amounts(grouped[ex.report_date], ex.amounts)

    if undated:
        print(f"WARN skipped undated PDFs: {undated}", file=sys.stderr)

    return grouped


def first_empty_column_by_row(ws, row_idx: int, start_col: int) -> int:
    col = start_col
    while True:
        v = ws.cell(row=row_idx, column=col).value
        if v is None or str(v).strip() == "":
            return col
        col += 1


def _translate_formula(formula: str, origin: str, target: str) -> str:
    try:
        return Translator(formula, origin=origin).translate_formula(target)
    except Exception:
        return formula


def copy_column(ws, src_col: int, dst_col: int) -> None:
    for r in range(1, ws.max_row + 1):
        src = ws.cell(row=r, column=src_col)
        dst = ws.cell(row=r, column=dst_col)

        value = src.value
        if isinstance(value, ArrayFormula):
            # Keep array-formula semantics, but rebuild with translated text/ref.
            formula_text = value.text if getattr(value, "text", None) else None
            if formula_text:
                if not formula_text.startswith("="):
                    formula_text = "=" + formula_text
                translated = _translate_formula(formula_text, src.coordinate, dst.coordinate)
                # ArrayFormula expects formula text (with or without '=') and its own ref.
                dst.value = ArrayFormula(ref=dst.coordinate, text=translated)
            else:
                dst.value = None
        elif src.data_type == "f" and isinstance(value, str) and value.startswith("="):
            dst.value = _translate_formula(value, src.coordinate, dst.coordinate)
        else:
            dst.value = value

        if src.has_style:
            dst._style = copy(src._style)
        if src.comment is not None:
            dst.comment = copy(src.comment)
        if src.hyperlink is not None:
            dst._hyperlink = copy(src.hyperlink)

    src_letter = get_column_letter(src_col)
    dst_letter = get_column_letter(dst_col)
    if src_letter in ws.column_dimensions:
        ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width


EXPENSE_CODES_AS_ABS = {2120, 2210, 2220, 2330, 2350}


def transform_2410_for_excel(raw_value: int, amounts: dict[int, int]) -> int:
    if raw_value == 0:
        return 0

    # If both 2300 and 2400 are present, infer whether 2410 reduces or increases net profit.
    # 2400 < 2300 => reducing effect => Excel value should be positive (formula subtracts 2410).
    # 2400 > 2300 => increasing effect => Excel value should be negative.
    before_tax = amounts.get(2300)
    net_profit = amounts.get(2400)
    if isinstance(before_tax, int) and isinstance(net_profit, int):
        if net_profit < before_tax:
            return abs(raw_value)
        if net_profit > before_tax:
            return -abs(raw_value)

    # Fallback: preserve source effect by inverting source sign.
    return -raw_value


def transform_amount_for_excel(code: int, raw_value: int, amounts: dict[int, int]) -> int:
    if code in EXPENSE_CODES_AS_ABS:
        return abs(raw_value)

    if code == 2410:
        return transform_2410_for_excel(raw_value, amounts)

    return raw_value


def fill_column_by_codes(ws, target_col: int, amounts: dict[int, int]) -> int:
    filled = 0
    for r in range(1, ws.max_row + 1):
        code = normalize_code(ws.cell(row=r, column=5).value)
        if code is None or code not in amounts:
            continue

        cell = ws.cell(row=r, column=target_col)
        if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
            continue

        raw_value = amounts[code]
        cell.value = transform_amount_for_excel(code, raw_value, amounts)
        filled += 1
    return filled


def build_transformed_amounts(amounts: dict[int, int]) -> dict[int, int]:
    return {
        code: transform_amount_for_excel(code, raw_value, amounts)
        for code, raw_value in amounts.items()
    }


def write_output_csv(output_xlsm: Path, date_to_amounts: dict[dt.date, dict[int, int]]) -> Path:
    csv_path = output_xlsm.with_suffix(".csv")
    ordered_dates = sorted(date_to_amounts)
    transformed_by_date = {
        d: build_transformed_amounts(date_to_amounts[d]) for d in ordered_dates
    }

    all_codes = sorted({
        code
        for one_date_amounts in transformed_by_date.values()
        for code in one_date_amounts.keys()
    })

    with csv_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, delimiter=";", lineterminator="\n")
        for code in all_codes:
            row = [str(code)]
            for d in ordered_dates:
                v = transformed_by_date[d].get(code)
                row.append("" if v is None else str(v))
            writer.writerow(row)

    return csv_path


def make_output_xlsm_name(input_xlsm: Path, actual_report_date: dt.date) -> Path:
    date_str = (actual_report_date + dt.timedelta(days=1)).strftime("%d.%m.%Y")
    stem = _repair_mojibake_name(input_xlsm.stem)

    def _valid_date_token(token: str, token_kind: str) -> bool:
        try:
            if token_kind == "compact":
                dt.datetime.strptime(token, "%d%m%Y")
                return True
            parts = re.split(r"[._\-]", token)
            if len(parts) != 3:
                return False
            if token_kind == "dmy":
                dt.date(int(parts[2]), int(parts[1]), int(parts[0]))
            else:
                dt.date(int(parts[0]), int(parts[1]), int(parts[2]))
            return True
        except Exception:
            return False

    patterns = [
        (re.compile(r"(?<!\d)(\d{2}[._\-]\d{2}[._\-]\d{4})(?!\d)"), "dmy"),
        (re.compile(r"(?<!\d)(\d{4}[._\-]\d{2}[._\-]\d{2})(?!\d)"), "ymd"),
        (re.compile(r"(?<!\d)(\d{8})(?!\d)"), "compact"),
    ]

    matches: list[tuple[int, int]] = []
    for regex, kind in patterns:
        for m in regex.finditer(stem):
            token = m.group(1)
            if _valid_date_token(token, kind):
                matches.append((m.start(1), m.end(1)))

    if matches:
        start, end = max(matches, key=lambda p: p[0])
        new_stem = stem[:start] + date_str + stem[end:]
        return input_xlsm.with_name(new_stem + input_xlsm.suffix)

    return input_xlsm.with_name(stem + "_out" + input_xlsm.suffix)


def apply_dates_to_xlsm(
    input_xlsm: Path,
    date_to_amounts: dict[dt.date, dict[int, int]],
    output_xlsm: Path,
) -> list[str]:
    wb = load_workbook(input_xlsm, keep_vba=True, data_only=False)
    ws = wb["ФО-ввод"]
    w_ctrl = wb["Управление"]

    logs: list[str] = []

    for report_date in sorted(date_to_amounts):
        target_col = first_empty_column_by_row(ws, row_idx=8, start_col=7)
        copy_column(ws, src_col=6, dst_col=target_col)

        header_date = report_date + dt.timedelta(days=1)
        head = ws.cell(row=8, column=target_col)
        head.value = header_date
        head.number_format = "DD.MM.YYYY"

        filled = fill_column_by_codes(ws, target_col, date_to_amounts[report_date])
        logs.append(
            f"date={report_date.isoformat()} -> col={get_column_letter(target_col)} filled={filled}"
        )

    latest_date = max(date_to_amounts)
    w_ctrl["D4"].value = latest_date + dt.timedelta(days=1)
    w_ctrl["D4"].number_format = "DD.MM.YYYY"

    wb.save(output_xlsm)
    return logs


def find_files(root: Path, exts: tuple[str, ...]) -> list[Path]:
    exts_lower = {e.lower() for e in exts}
    return sorted(
        p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in exts_lower and not p.name.startswith("~$")
    )


def find_direct_files(root: Path, exts: tuple[str, ...]) -> list[Path]:
    exts_lower = {e.lower() for e in exts}
    return sorted(
        p for p in root.iterdir() if p.is_file() and p.suffix.lower() in exts_lower and not p.name.startswith("~$")
    )


def get_default_input_root() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return SCRIPT_DIR


def detect_auto_scenario(input_path: Path) -> tuple[str, Path]:
    if input_path.is_file():
        if input_path.suffix.lower() == ".zip":
            return "B", input_path
        raise RuntimeError("AUTO: input file must be .zip, or pass a directory")

    if not input_path.is_dir():
        raise RuntimeError(f"AUTO: input path is neither file nor directory: {input_path}")

    direct_dirs = sorted([p for p in input_path.iterdir() if p.is_dir()])
    direct_zips = find_direct_files(input_path, (".zip",))
    direct_pdfs = find_direct_files(input_path, (".pdf",))
    direct_xlsms = find_direct_files(input_path, (".xlsm",))

    if (direct_dirs or direct_zips) and not direct_pdfs and not direct_xlsms:
        return "B", input_path

    if direct_pdfs or direct_xlsms:
        if direct_dirs or direct_zips:
            raise RuntimeError(
                "AUTO: mixed structure detected (subfolders/zip together with root PDF/XLSM). "
                "Use explicit scenario or separate inputs."
            )
        if not direct_pdfs:
            raise RuntimeError("AUTO: XLSM found, but no PDF files in input directory")
        if not direct_xlsms:
            raise RuntimeError("AUTO: PDF files found, but no XLSM in input directory")
        return "AUTO", input_path

    raise RuntimeError("AUTO: no input data found (expect zip, company subfolders, or PDF+XLSM set)")


def process_single_folder(root: Path, dpi: int, lang: str, psm: int, scenario: str, export_csv: bool = False) -> list[Path]:
    pdfs = find_files(root, (".pdf",))
    xlsms = find_files(root, (".xlsm",))

    if not pdfs:
        raise RuntimeError(f"{scenario}: no PDF files found in {root}")
    if not xlsms:
        raise RuntimeError(f"{scenario}: no XLSM files found in {root}")
    if len(xlsms) != 1:
        raise RuntimeError(f"{scenario}: expected exactly 1 XLSM, found {len(xlsms)} in {root}")

    xlsm = xlsms[0]
    grouped = group_pdfs_by_date(
        pdfs,
        dpi=dpi,
        lang=lang,
        psm=psm,
        infer_pair_date=scenario in {"A", "B", "AUTO"},
    )
    if not grouped:
        raise RuntimeError(f"{scenario}: no dated PDF data extracted in {root}")

    effective_scenario = scenario
    if scenario == "AUTO":
        effective_scenario = "C" if len(grouped) > 1 else "A"
        print(
            f"INFO AUTO {root.name}: detected scenario={effective_scenario} "
            f"(report_dates={len(grouped)})",
            file=sys.stderr,
        )

    if effective_scenario in {"A", "B"} and len(grouped) > 1:
        latest = max(grouped)
        print(
            f"WARN {effective_scenario} {root.name}: multiple report dates detected ({len(grouped)}), "
            f"using latest {latest.isoformat()}",
            file=sys.stderr,
        )
        grouped = {latest: grouped[latest]}

    latest_report_date = max(grouped)
    output_xlsm = make_output_xlsm_name(xlsm, latest_report_date)
    logs = apply_dates_to_xlsm(xlsm, grouped, output_xlsm)

    outputs = [output_xlsm]
    output_csv: Path | None = None
    if export_csv:
        output_csv = write_output_csv(output_xlsm, grouped)
        outputs.append(output_csv)

    print(f"OK {effective_scenario} {root}: {xlsm.name} -> {output_xlsm.name}")
    for item in logs:
        print(f"  {item}")
    if output_csv is not None:
        print(f"  csv={output_csv.name}")

    return outputs


def run_scenario_a(input_path: Path, dpi: int, lang: str, psm: int, export_csv: bool = False) -> list[Path]:
    return process_single_folder(input_path, dpi=dpi, lang=lang, psm=psm, scenario="A", export_csv=export_csv)


def _run_b_company_dirs(root: Path, dpi: int, lang: str, psm: int, export_csv: bool = False) -> list[Path]:
    outputs: list[Path] = []
    company_dirs = sorted([p for p in root.iterdir() if p.is_dir()])
    for company in company_dirs:
        has_pdf = bool(find_files(company, (".pdf",)))
        has_xlsm = bool(find_files(company, (".xlsm",)))
        if not (has_pdf and has_xlsm):
            print(f"INFO B {company}: skipped (no PDF+XLSM set)", file=sys.stderr)
            continue

        try:
            out = process_single_folder(company, dpi=dpi, lang=lang, psm=psm, scenario="B", export_csv=export_csv)
            outputs.extend(out)
        except Exception as e:
            print(f"FAIL B {company}: {e}", file=sys.stderr)
    return outputs


def run_scenario_b(input_path: Path, dpi: int, lang: str, psm: int, export_csv: bool = False) -> list[Path]:
    if input_path.is_file() and input_path.suffix.lower() == ".zip":
        with tempfile.TemporaryDirectory(prefix="afhd_b_zip_") as td:
            extracted_root = Path(td) / "src"
            extracted_root.mkdir(parents=True, exist_ok=True)
            extract_zip_preserve_ru_names(input_path, extracted_root)

            company_dirs = sorted([p for p in extracted_root.iterdir() if p.is_dir()])
            if not company_dirs:
                raise RuntimeError(f"B: no company subfolders found in archive {input_path}")

            temp_outputs = _run_b_company_dirs(extracted_root, dpi=dpi, lang=lang, psm=psm, export_csv=export_csv)
            if not temp_outputs:
                raise RuntimeError(f"B: no outputs generated from archive {input_path}")

            out_root = input_path.with_name(input_path.stem + "_out")
            if out_root.exists():
                shutil.rmtree(out_root)
            out_root.mkdir(parents=True, exist_ok=True)

            copied_outputs: list[Path] = []
            for src in temp_outputs:
                rel_parent = src.parent.relative_to(extracted_root)
                dst_dir = out_root / rel_parent
                dst_dir.mkdir(parents=True, exist_ok=True)
                dst = dst_dir / src.name
                shutil.copy2(src, dst)
                copied_outputs.append(dst)

            print(f"B zip outputs copied to: {out_root}")
            return copied_outputs

    if not input_path.is_dir():
        raise RuntimeError("B: input must be a directory or .zip")

    outputs: list[Path] = []
    company_dirs = sorted([p for p in input_path.iterdir() if p.is_dir()])
    zip_files = find_direct_files(input_path, (".zip",))

    if company_dirs:
        outputs.extend(_run_b_company_dirs(input_path, dpi=dpi, lang=lang, psm=psm, export_csv=export_csv))

    for zip_path in zip_files:
        try:
            outputs.extend(run_scenario_b(zip_path, dpi=dpi, lang=lang, psm=psm, export_csv=export_csv))
        except Exception as e:
            print(f"FAIL B {zip_path}: {e}", file=sys.stderr)

    if outputs:
        return outputs

    pdfs_here = find_direct_files(input_path, (".pdf",))
    xlsms_here = find_direct_files(input_path, (".xlsm",))
    if pdfs_here and xlsms_here:
        print(
            f"WARN B {input_path.name}: no company subfolders; treating directory as a single company set",
            file=sys.stderr,
        )
        return process_single_folder(input_path, dpi=dpi, lang=lang, psm=psm, scenario="B", export_csv=export_csv)

    raise RuntimeError(
        f"B: no processable input found in {input_path} "
        "(expected company subfolders, zip files, or local PDF+XLSM set)"
    )


def run_scenario_c(input_path: Path, dpi: int, lang: str, psm: int, export_csv: bool = False) -> list[Path]:
    return process_single_folder(input_path, dpi=dpi, lang=lang, psm=psm, scenario="C", export_csv=export_csv)


def run_scenario_auto(input_path: Path, dpi: int, lang: str, psm: int, export_csv: bool = False) -> list[Path]:
    if not input_path.is_dir():
        raise RuntimeError("AUTO (A/C): input must be a directory with PDF+XLSM set")
    return process_single_folder(input_path, dpi=dpi, lang=lang, psm=psm, scenario="AUTO", export_csv=export_csv)


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Transfer financial report data from PDF to XLSM templates")
    p.add_argument(
        "scenario",
        nargs="?",
        choices=["A", "B", "C", "AUTO"],
        default="AUTO",
        help="A=Single, B=Many, C=Multi-date, AUTO=detect by input structure",
    )
    p.add_argument(
        "input",
        nargs="?",
        type=Path,
        default=None,
        help="Input folder/file according to scenario (default: folder with script/exe)",
    )
    p.add_argument("--dpi", type=int, default=300, help="OCR DPI (default: 300)")
    p.add_argument("--lang", default="rus", help="Tesseract language (default: rus)")
    p.add_argument("--psm", type=int, default=4, help="Tesseract PSM (default: 4)")
    p.add_argument("--export-csv", action="store_true", help="Also export CSV next to output XLSM")
    return p


def main() -> int:
    args = build_parser().parse_args()
    input_path = args.input if args.input is not None else get_default_input_root()
    scenario = args.scenario

    if not input_path.exists():
        print(f"Input path not found: {input_path}", file=sys.stderr)
        return 2

    try:
        if scenario == "AUTO":
            scenario, input_path = detect_auto_scenario(input_path)
            if scenario == "AUTO":
                print(f"INFO AUTO: detected single-folder mode (A/C by report dates) -> {input_path}")
            else:
                print(f"INFO AUTO: detected scenario={scenario} -> {input_path}")

        if scenario == "A":
            outputs = run_scenario_a(input_path, dpi=args.dpi, lang=args.lang, psm=args.psm, export_csv=args.export_csv)
        elif scenario == "B":
            outputs = run_scenario_b(input_path, dpi=args.dpi, lang=args.lang, psm=args.psm, export_csv=args.export_csv)
        elif scenario == "C":
            outputs = run_scenario_c(input_path, dpi=args.dpi, lang=args.lang, psm=args.psm, export_csv=args.export_csv)
        else:
            outputs = run_scenario_auto(input_path, dpi=args.dpi, lang=args.lang, psm=args.psm, export_csv=args.export_csv)

        print(f"Done: generated {len(outputs)} file(s)")
        for out in outputs:
            print(f"  {out}")
        return 0
    except Exception as e:
        print(f"FAIL: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
